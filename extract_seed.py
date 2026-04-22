import openpyxl, json, datetime, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

SRC = '../2026分揀揀次.xlsx'
TODAY = datetime.date(2026, 4, 22)

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['總表']

date_col = {}
for c in range(3, ws.max_column + 1):
    v = ws.cell(1, c).value
    if hasattr(v, 'strftime'):
        date_col[v.strftime('%Y-%m-%d')] = c

STATIONS = ['台東','岡山1','花蓮倉','大肚1','岡山2','大肚2','大溪1','新光二倉1','大溪2','新光二倉2','大溪3']
A_ROWS = list(range(58, 69))   # 58..68
B_ROWS = list(range(72, 83))   # 72..82

standards = {}
for i, s in enumerate(STATIONS):
    a_t = ws.cell(A_ROWS[i], 1).value
    b_t = ws.cell(B_ROWS[i], 1).value
    standards[s] = {
        'A': a_t.strftime('%H:%M') if a_t else None,
        'B': b_t.strftime('%H:%M') if b_t else None,
    }

std_picks = ws.cell(55, 1).value
std_boxes = ws.cell(56, 1).value
std_start = ws.cell(57, 1).value
standards_meta = {
    'picks': std_picks,
    'boxes': std_boxes,
    'A_start': std_start.strftime('%H:%M') if std_start else None,
    'stations': standards,
}

def fmt_time(v):
    if v is None:
        return None
    if hasattr(v, 'strftime'):
        return v.strftime('%H:%M')
    return None

records = []
for i in range(30, 0, -1):
    d = TODAY - datetime.timedelta(days=i)
    key = d.strftime('%Y-%m-%d')
    if key not in date_col:
        continue
    c = date_col[key]
    total_picks = ws.cell(20, c).value
    total_boxes = ws.cell(32, c).value
    if not total_picks:
        continue
    a_stations = {}
    b_stations = {}
    for j, s in enumerate(STATIONS):
        a_stations[s] = fmt_time(ws.cell(A_ROWS[j], c).value)
        b_stations[s] = fmt_time(ws.cell(B_ROWS[j], c).value)
    def to_min(t):
        h, m = map(int, t.split(':'))
        mins = h * 60 + m
        if mins < 12 * 60:
            mins += 24 * 60
        return mins
    all_times = [t for t in list(a_stations.values()) + list(b_stations.values()) if t]
    if all_times:
        latest = max(all_times, key=to_min)
        total_end = latest
    else:
        total_end = None
    records.append({
        'date': key,
        'weekday': d.weekday(),
        'totalPicks': total_picks,
        'totalBoxes': total_boxes,
        'aBoxes': ws.cell(30, c).value,
        'bBoxes': ws.cell(31, c).value,
        'aStart': fmt_time(ws.cell(35, c).value),
        'bStart': fmt_time(ws.cell(36, c).value),
        'aStations': a_stations,
        'bStations': b_stations,
        'totalEnd': total_end,
    })

out = {
    'generatedAt': datetime.datetime.now().isoformat(timespec='seconds'),
    'today': TODAY.strftime('%Y-%m-%d'),
    'standards': standards_meta,
    'stations': STATIONS,
    'records': records,
}

js = 'window.SEED = ' + json.dumps(out, ensure_ascii=False, indent=2) + ';\n'
with open('seed-data.js', 'w', encoding='utf-8') as f:
    f.write(js)

print(f'Wrote {len(records)} records to seed-data.js')
print('Latest 3:', [r['date'] for r in records[-3:]])
print('Standards:', standards_meta)
